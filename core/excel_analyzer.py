import io
import openpyxl
from openpyxl.utils import get_column_letter

class ExcelAnalyzer:
    def __init__(self, file: bytes):
        self.__workbook: openpyxl.Workbook = openpyxl.load_workbook(io.BytesIO(file), read_only=True)

    def get_cell_values(self, sheet_name: str, column_index: int, offset: int = 1) -> list[str]:
        sheet = self.__workbook[sheet_name]

        values: list[str] = []

        for row in sheet.iter_rows(min_row=offset, values_only=True):
            if len(row) >= column_index - 1:
                value = row[column_index - 1]

                if value is not None:
                    values.append(value)

        return values

    @property
    def sheet_names(self) -> list[str]:
        return self.__workbook.sheetnames

    def get_active_columns(self, sheet_name) -> list[str]:
        sheet = self.__workbook[sheet_name]

        active_columns: set[int] = set()

        for row in sheet.iter_rows(values_only=True):
            for column_index, cell in enumerate(row, start=1):
                if cell is not None and cell != "":
                    active_columns.add(column_index)

        return [get_column_letter(index) for index in sorted(active_columns)]

    def __del__(self):
        self.__workbook.close()